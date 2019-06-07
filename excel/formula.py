#! python3

import openpyxl, os

import setup

filename = 'writeFormula.xlsx'

# create
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)'
wb.save(filename)


wb_formulas = openpyxl.load_workbook(filename, data_only=False)
sheet = wb_formulas.active
formula = sheet['A3'].value

wb_data_only = openpyxl.load_workbook(filename, data_only=True)
sheet = wb_data_only.active
value = sheet['A3'].value

print('{}, {}'.format(formula, value)) # => =SUM(A1:A2), None

