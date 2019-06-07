#! python3

import openpyxl
import setup

# create data cell
wb = openpyxl.Workbook()
for i in range(1, 11):
    wb.active.cell(i, 1).value = i * i
    wb.active.cell(i, 2).value = i

# draw bar chart
ref_obj = openpyxl.chart.Reference(wb.active, min_col=1, min_row=1, max_col=1, max_row=10)
series_obj = openpyxl.chart.Series(ref_obj, title='First series')
chart_obj = openpyxl.chart.BarChart()

chart_obj.append(series_obj)
chart_obj.y = 50
chart_obj.x = 100
chart_obj.w = 300
chart_obj.h = 200

wb.active.add_chart(chart_obj)

# TODO: no draw
ref_obj = openpyxl.chart.Reference(wb.active, min_col=2, min_row=1, max_col=2, max_row=10)
openpyxl.Workbook().active.add_chart(
    openpyxl.chart.BarChart().append(
        openpyxl.chart.Series(ref_obj, title='Second series')
    )
)

wb.save('sampleChart.xlsx')

