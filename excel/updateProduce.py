#! python3
# updateProduce.py -

import openpyxl
import setup

wb = openpyxl.load_workbook('/tmp/produceSales.xlsx')
sheet = wb['Sheet']

PRICE_UPDATE = {'Garlic': 3.07,
                'Celery': 1.19,
                'Lemon': 1.27}


for row_num in range(2, sheet.max_row + 1): # skip first row
    produce_name = sheet.cell(row=row_num, column=1).value
    if produce_name in PRICE_UPDATE:
        sheet.cell(row=row_num, column=2).value = PRICE_UPDATE[produce_name]


wb.save('produceSales_updated.xlsx')
