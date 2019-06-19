#! python3
# check spreadsheet paid status. unpaid users send Email

import openpyxl, smtplib, sys

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb['Sheet1']

last_col = sheet.get_highest_column()

latest_month = sheet.cell(row=1, column=last_col).value

unpaid_members = {}
for r in range(2, sheet.get_highest_row() + 1):
    payment = sheet.cell(row=r, column=last_col).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaid_members[name] = email

smtp_obj = smtplib.SMTP('smtp.gmail.com', 587)
smtp_obj.ehlo()
smtp_obj.starttls()
smtp_obj.login('hoge@gmail.com', 'huga')

for name, email in unpaid_members.items():
    body = """Subject: {} dues unpaid.
    Dear {},
    Records show that you have not paid dues for {}.
    Please make this payment as soon as possible.
    Thank you!""".format(latest_month, name, latest_month)

    sendmail_status = smtp_obj.sendmail('no-reply@gmail.com', email, body)
    if sendmail_status != {}:
        print('Sending Error to {}: {}'.format(email, sendmail_status))

smtp_obj.quit()

